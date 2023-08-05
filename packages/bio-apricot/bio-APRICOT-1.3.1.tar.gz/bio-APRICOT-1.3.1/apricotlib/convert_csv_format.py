#! /usr/bin/env python
# Description = This script is designed to convert the file types.
# tab separated files to xlsx files

import os
import sys
import csv
import shutil
csv.field_size_limit(sys.maxsize)
try:
    from openpyxl.workbook import Workbook
except ImportError:
    print('Python package openpyxl is missing. Please install/update.\n'
          'Please ignore if you chose the output format as HTML')
def csv_to_xlsx(inpath, outpath):
    '''Converts table to excel'''
    allowed_format_list = ['csv', 'tab', 'tsv']
    try:
        for csv_file in os.listdir(inpath):
            if os.path.isfile(inpath+'/'+csv_file):
                if csv_file.split('.')[1] in allowed_format_list:
                    sheet_name = csv_file.split('.')[0]
                    workbook = Workbook()
                    xlsxsheet = workbook.create_sheet(0)
                    for i, row in enumerate(
                            csv.reader(open(inpath+'/'+csv_file), delimiter="\t"), 1):
                        for j, col in enumerate(row, 1):
                            xlsxsheet.cell(row=i, column=j).value = col
                    workbook.save(open(outpath+'/'+sheet_name+'.xlsx', 'wb'))
                else:
                    print('%s: This text file will not be converted into the excel format!' % csv_file)
                    shutil.copyfile(inpath+'/'+csv_file, outpath+'/'+csv_file)
            elif os.path.isdir(inpath+'/'+csv_file):
                if not os.path.exists(outpath+'/'+csv_file):
                    os.mkdir(outpath+'/'+csv_file)
                csv_to_xlsx(inpath+'/'+csv_file, outpath+'/'+csv_file)
    except NotADirectoryError:
        pass

def csv_to_html(inpath, outpath):
    '''Convert table to HTML'''
    allowed_format_list = ['csv', 'tab', 'tsv', 'txt']
    try:
        for csv_file in os.listdir(inpath):
            if os.path.isfile(inpath+'/'+csv_file):
                if csv_file.split('.')[1] in allowed_format_list:
                    with open(outpath+'/'+csv_file.split('.')[0]+'.html',"w") as outfile:
                        outfile.write('\n'.join(['<!DOCTYPE html>', '<html lang="en">', '<head>',
                        '<meta charset="utf-8">',
                        '<link rel="stylesheet" type="text/css" href="http://ajax.aspnetcdn.com/ajax/jquery.dataTables/1.9.4/css/jquery.dataTables.css">',
                        '<script type="text/javascript" charset="utf8" src="http://ajax.aspnetcdn.com/ajax/jQuery/jquery-1.8.2.min.js"></script>',
                        '<script type="text/javascript" charset="utf8" src="http://ajax.aspnetcdn.com/ajax/jquery.dataTables/1.9.4/jquery.dataTables.min.js"></script>',
                        '<script type="text/javascript" language="javascript" class="init">',
                        '$(document).ready(function () {',
                        '\t$("#subscriptionlist").dataTable();',
                        '});', '</script>', '<style type="text/CSS">',
                        'table, th, td {', '\tborder: 1px solid #E8E8E8;',
                        '\tborder-collapse: collapse;', '\tmin-width:150px;',
                        '\ttext-align:left', '}', '</style>',
                        '</head>', '<div id="subscriptionsList">',
                        '<h2>Source file: %s</h2>' % csv_file,
                        '<h3>Source path: %s</h3>' % inpath, '<table id="subscriptionlist">\n']))
                        row_num = 0
                        table_string = ""
                        if os.path.getsize(inpath+'/'+csv_file) > 0:
                            if csv_file.split('.')[1] == 'txt':
                                for lines in open(inpath+'/'+csv_file).readlines():
                                    if not '\t' in lines:
                                        outfile.write("<pre>%s</pre>" % lines)
                                    else:
                                        table_style = """<style>
                                        table {border-collapse:collapse; table-layout:fixed; width:310px;}
                                        table td {border:solid 1px #fab; width:300px; word-wrap:break-word;}
                                        </style>"""
                                        table_entry = '%s%s%s%s' % (
                                            '    <tr><th>','</th><th>'.join(
                                                lines.strip().split('\t')),'</td></tr>','\n')
                                        outfile.write('%s%s%s%s' % ('<table>\n',
                                                    table_style, table_entry, '</table>'))
                            else:
                                csv_fh = csv.reader(open(inpath+'/'+csv_file), delimiter="\t")
                                for i, row in enumerate(csv_fh):
                                    if 'annotation_scoring' in csv_file:
                                        if 'annotation_scoring_of_selected_data_filter' in csv_file:
                                            for idx in range(1, 3):
                                                if len(list(row[idx])) > 50:
                                                    row[idx] = '<br>'.join(
                                                        split_str(row[idx], 50))
                                        else:
                                            for idx in range(13, 17):
                                                if len(list(row[idx])) > 50:
                                                    row[idx] = '<br>'.join(
                                                        split_str(row[idx], 50))
                                    if i == 0:
                                        outfile.write("<thead>\n<tr>\n<th>")
                                        outfile.write('</th><th>'.join(row)+'\n')
                                        outfile.write("</th>\n</tr>\n</thead>\n<tbody>\n")
                                    else:
                                        outfile.write("<tr>\n<td>")
                                        outfile.write('</td><td>'.join(row)+'\n')
                                        outfile.write("</td>\n</tr>\n")
                        outfile.write(
                            '\n'.join(['</tbody>', '</table>', '</div>', '</html>']))
                else:
                    print('This file can not be converted into the html format: %s' % csv_file)
                    shutil.copyfile(inpath+'/'+csv_file, outpath+'/'+csv_file)
                    
            elif os.path.isdir(inpath+'/'+csv_file):
                if not os.path.exists(outpath+'/'+csv_file):
                    os.mkdir(outpath+'/'+csv_file)
                csv_to_html(inpath+'/'+csv_file, outpath+'/'+csv_file)
    except NotADirectoryError:
        pass

def split_str(seq, chunk, skip_tail=False):
    lst = []
    if chunk <= len(seq):
        lst.extend([seq[:chunk]])
        lst.extend(split_str(seq[chunk:], chunk, skip_tail))
    elif not skip_tail and seq:
        lst.extend([seq])
    return lst
