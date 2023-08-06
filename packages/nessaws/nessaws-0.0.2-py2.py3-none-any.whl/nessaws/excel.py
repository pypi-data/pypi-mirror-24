"""Write results to Excel file."""
import csv
import datetime
import logging
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.writer.write_only import WriteOnlyCell


logger = logging.getLogger('nessaws.excel')


def write_excel_output(results):
    """Write results from Nessus scans to a single Excel workbook.

    Args:
        results (dict): The `results` dictionary after Nessus scans have
            been completed.

    Returns:
        A string of the filename exported by this function.

    """
    # create workbook and summary sheet
    output_workbook = Workbook(write_only=True)
    summary_sheet = output_workbook.create_sheet('Summary')
    title_cell = WriteOnlyCell(summary_sheet, value='NessAWS Scan Results')
    title_cell.font = Font(size=18, bold=True)
    summary_sheet.append([title_cell])
    summary_sheet.append([])

    scan_accounts_title = WriteOnlyCell(summary_sheet, value='AWS Accounts')
    scan_accounts_title.font = Font(size=12, bold=True)
    scan_accounts = WriteOnlyCell(summary_sheet, value=', '.join(
        results['account_names']))
    summary_sheet.append([scan_accounts_title, scan_accounts])
    tag_values_title = WriteOnlyCell(summary_sheet, value='Tag Values')
    tag_values_title.font = Font(size=12, bold=True)
    tag_values = WriteOnlyCell(summary_sheet, value=', '.join(
        results['tag_values']))
    summary_sheet.append([tag_values_title, tag_values])
    start_date_title = WriteOnlyCell(summary_sheet, value='Start Date')
    start_date_title.font = Font(size=12, bold=True)
    start_date = WriteOnlyCell(summary_sheet, value=results['start_date'])
    summary_sheet.append([start_date_title, start_date])
    end_date_title = WriteOnlyCell(summary_sheet, value='End Date')
    end_date_title.font = Font(size=12, bold=True)
    end_date = WriteOnlyCell(summary_sheet, value=results['end_date'])
    summary_sheet.append([end_date_title, end_date])
    summary_sheet.append([])

    scan_names = WriteOnlyCell(summary_sheet, value='Nessus Scan Name')
    scan_names.font = Font(size=12, bold=True)
    status = WriteOnlyCell(summary_sheet, value='Status')
    status.font = Font(size=12, bold=True)
    targets = WriteOnlyCell(summary_sheet, value='Targets')
    targets.font = Font(size=12, bold=True)
    credentialed_targets = WriteOnlyCell(
        summary_sheet, value='Credentialed Targets')
    credentialed_targets.font = Font(size=12, bold=True)
    uncredentialed_targets = WriteOnlyCell(
        summary_sheet, value='Uncredentialed Targets')
    uncredentialed_targets.font = Font(size=12, bold=True)
    summary_sheet.append(
        [scan_names, status, targets, credentialed_targets,
         uncredentialed_targets])

    # update summary sheet with data on scans
    for scan in results['scans']:
        scan_name = WriteOnlyCell(summary_sheet, value=scan['scan_name'])
        scan_status = WriteOnlyCell(summary_sheet, value=scan['status'])
        scan_targets = WriteOnlyCell(
            summary_sheet, value=', '.join(scan['target_names']))
        scan_credentialed_targets = WriteOnlyCell(
            summary_sheet, value=len(scan.get('credentialed_hosts', [])))
        scan_uncredentialed_targets = WriteOnlyCell(
            summary_sheet, value=len(scan.get('uncredentialed_hosts', [])))
        summary_sheet.append(
            [scan_name, scan_status, scan_targets, scan_credentialed_targets,
             scan_uncredentialed_targets])

    # create results sheet for all Nessus CSVs
    results_sheet = output_workbook.create_sheet('Results')

    plugin_id = WriteOnlyCell(results_sheet, value='Plugin ID')
    plugin_id.font = Font(size=12, bold=True)
    cve = WriteOnlyCell(results_sheet, value='CVE')
    cve.font = Font(size=12, bold=True)
    cvss = WriteOnlyCell(results_sheet, value='CVSS')
    cvss.font = Font(size=12, bold=True)
    risk = WriteOnlyCell(results_sheet, value='Risk')
    risk.font = Font(size=12, bold=True)
    host = WriteOnlyCell(results_sheet, value='Host')
    host.font = Font(size=12, bold=True)
    protocol = WriteOnlyCell(results_sheet, value='Protocol')
    protocol.font = Font(size=12, bold=True)
    port = WriteOnlyCell(results_sheet, value='Port')
    port.font = Font(size=12, bold=True)
    name = WriteOnlyCell(results_sheet, value='Name')
    name.font = Font(size=12, bold=True)
    synopsis = WriteOnlyCell(results_sheet, value='Synopsis')
    synopsis.font = Font(size=12, bold=True)
    description = WriteOnlyCell(results_sheet, value='Description')
    description.font = Font(size=12, bold=True)
    solution = WriteOnlyCell(results_sheet, value='Solution')
    solution.font = Font(size=12, bold=True)
    see_also = WriteOnlyCell(results_sheet, value='See Also')
    see_also.font = Font(size=12, bold=True)
    plugin_output = WriteOnlyCell(results_sheet, value='Plugin Output')
    plugin_output.font = Font(size=12, bold=True)
    results_sheet.append(
        [plugin_id, cve, cvss, risk, host, protocol, port, name, synopsis,
         description, solution, see_also, plugin_output])

    # loop through the scans and parse each CSV result file
    for scan in results['scans']:
        if scan['status'] == 'completed' and scan.get(
                'result_file') is not None:
            result_filename = scan['result_file']
            try:
                with open(result_filename) as csv_file:
                    reader = csv.reader(csv_file)
                    # skip the header row for each CSV
                    next(reader)
                    for row in reader:
                        data_row = []
                        for c, column in enumerate(row):
                            # change color on 'Risk' field
                            fill = None
                            if c == 3:
                                if column == 'Critical':
                                    fill = PatternFill(
                                        start_color='FF0000',
                                        end_color='FF0000',
                                        fill_type='solid')
                                if column == 'High':
                                    fill = PatternFill(
                                        start_color='FFA500',
                                        end_color='FFA500',
                                        fill_type='solid')
                                if column == 'Medium':
                                    fill = PatternFill(
                                        start_color='FFFF00',
                                        end_color='FFFF00',
                                        fill_type='solid')
                                if column == 'Low':
                                    fill = PatternFill(
                                        start_color='008000',
                                        end_color='008000',
                                        fill_type='solid')
                            # add AWS instance ID or name tag in the 'Host'
                            # field
                            if c == 4:
                                # if the 'Host' column matches either the
                                # public IP, private IP, or endpoint, add the
                                # instance name or ID to the column
                                for target in scan['targets']:
                                    if (str(target.get(
                                            'public_ip')) == column or str(
                                            target.get(
                                                'private_ip')) == column or
                                            target.get('endpoint') == column):
                                        column = '{}\n({})'.format(
                                            target['name'] if target.get(
                                                'name') else target['id'],
                                            column)
                                        break
                            data_cell = WriteOnlyCell(
                                results_sheet, value=column)
                            if fill:
                                data_cell.fill = fill
                            data_row.append(data_cell)
                        results_sheet.append(data_row)
                os.remove(result_filename)
            except (OSError, IOError):
                logger.warning(
                    'Could not find result file "{}" for scan name "{}", '
                    'these results will not be included in the report'.format(
                        result_filename, scan['scan_name']))

    # enable autofilter and save
    results_sheet.auto_filter.ref = 'A1:M1'
    output_filename = 'nessaws_{}.xlsx'.format(
        datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%s'))
    output_workbook.save(output_filename)
    return output_filename
