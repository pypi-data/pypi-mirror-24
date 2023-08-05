from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side
from nba_stats_api.constants import (GENERAL_STAT_TYPES,
                                     PLAYTYPE_COLUMNS,
                                     ALL_PLAY_TYPES,
                                     VIDEO_MEASURES,
                                     VIDEO_ENDPOINT)


class ExcelGenerator:
    def __init__(self, player_stats, season):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.current_column = None
        self.current_row = None
        self.video_row = 31
        self.video_column = 17
        self.player_stats = player_stats
        self.season = season

        self.header_font = Font(name="Calibri",
                                size=14,
                                bold="True")
        self.header_fill = PatternFill(start_color="ff8080",
                                       end_color="ff8080",
                                       fill_type="solid")
        self.row_fill = PatternFill(start_color="b3d9ff",
                                    end_color="b3d9ff",
                                    fill_type="solid")
        self.all_borders = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

    def _set_current_cell_value(self, value, row=None, column=None, **kwargs):
        self.worksheet.cell(row=row or self.current_row,
                            column=column or self.current_column,
                            value=value)
        for attr in kwargs:
            setattr(self.worksheet.cell(row=row or self.current_row, column=column or self.current_column),
                    attr, kwargs.get(attr))

    def _general_stats(self):
        for stat_type in GENERAL_STAT_TYPES:
            self._set_current_cell_value(value=stat_type,
                                         font=self.header_font)
            self.current_row += 1
            if self.player_stats.get(stat_type) is not None:
                for header_column in self.player_stats[stat_type]:
                    self._set_current_cell_value(value=header_column,
                                                 font=self.header_font,
                                                 fill=self.header_fill,
                                                 border=self.all_borders)
                    value = self.player_stats[stat_type][header_column]
                    self._set_current_cell_value(value=value,
                                                 row=self.current_row + 1,
                                                 fill=self.row_fill,
                                                 border=self.all_borders)
                    self.current_column += 1
            self.current_row += 3
            self.current_column = 1
        self.current_row += 2

    def _play_types(self):
        self._set_current_cell_value(value="Play Type",
                                     font=self.header_font,
                                     fill=self.header_fill,
                                     border=self.all_borders)
        self.current_column += 1
        for column in PLAYTYPE_COLUMNS:
            self._set_current_cell_value(value=column,
                                         font=self.header_font,
                                         fill=self.header_fill,
                                         border=self.all_borders)
            self.current_column += 1

        self.current_row += 1
        self.current_column = 1
        self.video_row = self.current_row

        for play_type in ALL_PLAY_TYPES:
            self._set_current_cell_value(value=play_type,
                                         font=self.header_font,
                                         fill=self.header_fill,
                                         border=self.all_borders)
            self.current_column += 1
            for column in self.player_stats.get(play_type, []):
                self._set_current_cell_value(value=self.player_stats[play_type][column],
                                             fill=self.row_fill,
                                             border=self.all_borders)
                self.current_column += 1
                self.video_column = self.current_column
            self.current_column = 1
            self.current_row += 1

        self.video_column += 1

    def _video(self):
        self.current_row = self.video_row
        self.current_column = self.video_column
        self._set_current_cell_value(value="Video (NBA.com archive)",
                                     font=self.header_font,
                                     fill=self.header_fill)
        self.current_row += 1

        for measure in VIDEO_MEASURES:
            url = VIDEO_ENDPOINT.format(player_id=self.player_stats['BasicInfo']['PLAYER_ID'],
                                        measure=measure, season=self.season,
                                        season_type="Regular+Season",
                                        )
            link_background_fill = PatternFill(start_color="1aff8c",
                                               end_color="1aff8c",
                                               fill_type="solid")
            self._set_current_cell_value(value=VIDEO_MEASURES[measure],
                                         fill=link_background_fill)
            self.worksheet.cell(row=self.current_row,
                                column=self.current_column).hyperlink = url
            self.current_row += 1

    def generate_workbook(self, auto_save=True, path="outputs"):
        player_name = self.player_stats['BasicInfo']['PLAYER_NAME']
        print(f"Creating Excel Spreadsheet for {player_name}. Autosave: {auto_save}")
        self.current_row = 1
        self.current_column = 1
        header_value = (self.player_stats['BasicInfo']['PLAYER_NAME'] +
                        " {season} stats and video hub".format(season=self.season))
        self._set_current_cell_value(value=header_value,
                                     font=self.header_font)
        self.current_row = 4

        self._general_stats()
        self._play_types()
        self._video()

        if auto_save:
            self.workbook.save(f"{path}/{player_name}.xlsx")

